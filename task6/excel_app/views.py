from rest_framework import viewsets
from openpyxl import load_workbook
from .models import Product
from .serializers import ProductSerializer
from rest_framework.response import Response
from rest_framework import status


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    queryset = Product.objects.all()
    
    def create(self, request, *args, **kwargs):
        try:
            file = request.data.get('file')
            workbook = load_workbook(file)
            ws = workbook.active
            headers = [cell.value for cell in ws[1]]
            data_from_excel = []
            for row in ws.iter_rows(2, values_only=True):
                data = dict(zip(headers, row))
                data['product_expiry_date']= data['product_expiry_date'].date()
                data['product_manufacturing_date']= data['product_manufacturing_date'].date()
                data_from_excel.append(data)
            serializer = self.get_serializer(data=data_from_excel, many=True)
            serializer.is_valid(raise_exception=True)
            created_products = []
            for data in data_from_excel:
                obj, created = Product.objects.get_or_create(**data)
                if created:
                    created_products.append(obj)
            serializer = ProductSerializer(created_products, many=True)
            return Response(data=serializer.data, status=201)
        except Exception as e:
            print(e)
            return Response(data={'detail':'Error Saving the data'}, status=400)
        









